from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import numpy as np
import cPickle

def one_hot(x,n):
    if type(x) == list:
	x = np.array(x)
    x = x.flatten()
    o_h = np.zeros((len(x),n))
    o_h[np.arange(len(x)),x] = 1
    return o_h

wb = load_workbook(filename = 'dataset.xlsx')
first_sheet = wb.get_sheet_names()[0]
ws = wb.get_sheet_by_name(first_sheet)

X = []
Y = []

for row in ws.iter_rows():
    r = map(lambda x: x.value, row)
    X.append(r[:3])
    Y.append(r[3])

X = X[1:]
Y = Y[1:]
Y = [y if y == 1 else 0 for y in Y]

X = np.array(X)
Y = one_hot(Y,2)

data = {"data" : X, "labels" : Y}

with open("data.pkl","wb") as f:
    cPickle.dump(data, f, -1)



