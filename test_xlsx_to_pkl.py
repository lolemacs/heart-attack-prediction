from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import numpy as np
import sys
import cPickle

wb = load_workbook(filename = sys.argv[1])
first_sheet = wb.get_sheet_names()[0]
ws = wb.get_sheet_by_name(first_sheet)

X = []

for row in ws.iter_rows():
    r = map(lambda x: x.value, row)
    X.append(r[:3])

X = X[1:]

X = np.array(X)

data = {"data" : X}

with open("test.pkl","wb") as f:
    cPickle.dump(data, f, -1)
